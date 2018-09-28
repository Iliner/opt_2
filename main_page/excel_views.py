import os

from django.shortcuts import redirect
from django.core.paginator import Paginator, InvalidPage
from .models import *
from .forms import *
from basket.models import Cart
from basket.forms import CartItemCount
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from main_page.twviews import *

import xlrd 
import xlwt
import pyexcel
import openpyxl

import smtplib
from email.mime.text import MIMEText
from email.header import Header



class Excel:
	"""Класс который делает список из файли xls
	
	>>>a = Excel('./excel/каталог_9.xls', code=0, producer=1, articul=2, photo=3)
	>>>b = Excel('./excel/фото_с_сервера.xls', code=0, producer=1, photo=3)

	"""
	def __init__(self, file_name, **kwargs):
		self.file = file_name
		self.name_column = kwargs
		self.common_list = None

	def create_list(self):
		rb = xlrd.open_workbook(self.file, formatting_info=True)
		sheet = rb.sheet_by_index(0)
		count = 0
		common_list = []
		for rownum in range(sheet.nrows):
		    row = sheet.row_values(rownum)
		    if self.name_column.get('code'):
		    	row[self.name_column['code']] = str(row[self.name_column['code']]).split('.')[0]
		    common_list.append(row)

		self.common_list = common_list
		return common_list

	def write_excel(self, file):
		file_for_writer = xlwt.Workbook()
		add_list_with_table = file_for_writer.add_sheet('Result')
		row = 0
		coll = 0
		while row < len(self.common_list):
			while coll < len(self.common_list[row]):
				add_list_with_table.write(row, coll, self.common_list[row][coll])
				coll += 1
			coll = 0  
			row +=1
		file_for_writer.save(file)

	
	def create_list_pyexcel(self):
		""" Создает массив

		Метож работает с помощью библиотеки pyexcel. 
		Форматы:
			-xls
			-xlsx
		"""
		self.common_list = pyexcel.get_array(file_name=self.file)
		return self.common_list

	def write_excel_pyexcel(self, file):
		""" Сохроняет в новый excel

		Форматы:
			-xls
			-xlsx
		"""

		pyexcel.save_as(array=self.common_list, dest_file_name=file)


	def write_excel_openpyxl(self, file):
		wb = openpyxl.load_workbook(filename = file)
		sheet = wb['TDSheet']
		row = 0
		collumn = 0
		lenght = len(self.common_list)
		while row < lenght:
			while collumn < len(self.common_list[row]):
				sheet.cell(row=row + 1, column=collumn + 1).value = self.common_list[row][collumn]
				collumn += 1
			row += 1
			collumn = 0	
		wb.save(file)



















@csrf_exempt
def working_excel(request):
	try:
		excel = ExcelImport.objects.all().filter(check=False).first()
		if excel:
			name = str(excel.file)
			BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
			full_path = "{}/uploads/{}".format(BASE_DIR, name)	
			print(full_path)
			object_excel = Excel(
				full_path, 
				code=1,
				producer=2,
				articul=3,
				description=4,
				stock=5,
				opt_6=6,
				opt_5=7, 
				opt_3=8,
				opt_2=9,
				opt_1=10
				)
			
			object_excel.create_list_pyexcel()
			for prosition in object_excel.common_list:
				code = prosition[object_excel.name_column['code']]
				check = Goods.objects.filter(code=code).exists()
				print(check)
				if not check:
					new_good = Goods()
					new_good.code = code
					new_good.articul = prosition[object_excel.name_column['articul']]
					check_producer = Producers.objects.filter(name=prosition[object_excel.name_column['producer']]).exists()
					if check_producer: 
						new_good.producer = Producers.objects.get(name=prosition[object_excel.name_column['producer']])
					else:
						new_producer = create_producer(prosition[object_excel.name_column['producer']]) 
						new_good.producer = new_producer
					new_good.description = prosition[object_excel.name_column['description']]
					new_good.in_stock = prosition[object_excel.name_column['stock']]
					new_good.price = prosition[object_excel.name_column['opt_1']]
					new_good.price_2 = prosition[object_excel.name_column['opt_2']]
					new_good.price_3 = prosition[object_excel.name_column['opt_3']]
					new_good.price_5 = prosition[object_excel.name_column['opt_5']]
					new_good.price_6 = prosition[object_excel.name_column['opt_10']]
					new_good.save()
					new_position = NewGoods()
					new_position.position = new_good
					new_position.save()
				else:
					pass
			
			excel.check = True
			excel.save()
		else:
			print('нет новых экселей')
	except Exception as err:
		print('my print' ,err)
		
		smtp_host = "smtp.mail.ru"
		smtp_port = "465"
		smtp_login = "stock@kvam.ru"
		smtp_password = "AT3TeC&5lshf"
		send_to = "ivan_1995i@mail.ru"

		message_text = str(err)

		smtp_send(smtp_host, smtp_port, smtp_login, smtp_password, send_to, message_text)






	response = 'hello'
	return HttpResponse(response)



def create_producer(name):
	new_producer = Producers()
	new_producer.name = name
	new_producer.save()
	return new_producer




def smtp_send(smtp_host, smtp_port, smtp_login, smtp_password, send_to, message_text):
	msg = MIMEText(message_text, 'plain', 'utf-8')
	msg['Subject'] = Header("Ошибка опт-тулс", 'utf-8')
	msg['From'] = smtp_login
	msg["To"] = send_to
	smtpObj = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=10)
	smtpObj.login(smtp_login, smtp_password)
	smtpObj.sendmail(smtp_login, send_to, msg.as_string())
	smtpObj.quit()




# def check_db(code):
# 	check = Goods.objects.filter(code=code).exists()
# 	if check:
# 		new_good = Goods()
# 		new_good.code = code
# 		new_good.articul = 
# 	else: