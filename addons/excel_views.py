import os
import random
import csv

from django.shortcuts import redirect
from django.core.paginator import Paginator, InvalidPage
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.base import View

from main_page.twviews import *
from .models import *
from filter.models import *
from basket.models import Cart
from basket.forms import CartItemCount

import xlrd 
import xlwt
import pyexcel
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa









class Excel:
	"""Класс который делает список из файли xls
	
	>>>a = Excel('./excel/каталог_9.xls', code=0, producer=1, articul=2, photo=3)
	>>>b = Excel('./excel/фото_с_сервера.xls', code=0, producer=1, photo=3)

	"""
	def __init__(self, file_name, **kwargs):
		self.file = file_name
		self.name_column = kwargs
		self.common_list = None

	def create_list(self):
		rb = xlrd.open_workbook(self.file, formatting_info=True)
		sheet = rb.sheet_by_index(0)
		count = 0
		common_list = []
		for rownum in range(sheet.nrows):
		    row = sheet.row_values(rownum)
		    if self.name_column.get('code'):
		    	row[self.name_column['code']] = str(row[self.name_column['code']]).split('.')[0]
		    common_list.append(row)

		self.common_list = common_list
		return common_list

	def write_excel(self, file):
		file_for_writer = xlwt.Workbook()
		add_list_with_table = file_for_writer.add_sheet('Result')
		row = 0
		coll = 0
		while row < len(self.common_list):
			while coll < len(self.common_list[row]):
				add_list_with_table.write(row, coll, self.common_list[row][coll])
				coll += 1
			coll = 0  
			row +=1
		file_for_writer.save(file)

	
	def create_list_pyexcel(self):
		""" Создает массив

		Метож работает с помощью библиотеки pyexcel. 
		Форматы:
			-xls
			-xlsx
		"""
		self.common_list = pyexcel.get_array(file_name=self.file)
		return self.common_list

	def write_excel_pyexcel(self, file):
		""" Сохроняет в новый excel

		Форматы:
			-xls
			-xlsx
		"""

		pyexcel.save_as(array=self.common_list, dest_file_name=file)


	def write_excel_openpyxl(self, file):
		wb = openpyxl.load_workbook(filename = file)
		sheet = wb['TDSheet']
		row = 0
		collumn = 0
		lenght = len(self.common_list)
		while row < lenght:
			while collumn < len(self.common_list[row]):
				sheet.cell(row=row + 1, column=collumn + 1).value = self.common_list[row][collumn]
				collumn += 1
			row += 1
			collumn = 0	
		wb.save(file)



















@csrf_exempt
def working_excel(request):
	# try:
	excel = ExcelImport.objects.all().filter(check=False).first()
	if excel:
		name = str(excel.file)
		BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
		full_path = "{}/uploads/{}".format(BASE_DIR, name)	
		meta = {}

		meta['code'] = excel.code
		if excel.producer:
			meta['producer'] = excel.producer
		if excel.articul:
			meta['articul'] = excel.articul
		if excel.description:
			meta['description'] = excel.description
		if excel.stock:
			meta['stock'] = excel.stock
		if excel.price:
			meta['price'] = excel.price
		if excel.opt_1:
			meta['opt_1'] = excel.opt_1
		if excel.opt_2:
			meta['opt_2'] = excel.opt_2
		if excel.opt_3:
			meta['opt_3'] = excel.opt_3
		if excel.opt_4:
			meta['opt_4'] = excel.opt_4
		if excel.opt_5:
			meta['opt_5'] = excel.opt_5
		if excel.opt_6:
			meta['opt_6'] = excel.opt_6
		if excel.opt_7:
			meta['opt_7'] = excel.opt_7
		if excel.opt_8:
			meta['opt_8'] = excel.opt_8
		if excel.opt_9:
			meta['opt_9'] = excel.opt_9
		if excel.opt_10:
			meta['opt_10'] = excel.opt_10
		if excel.opt_11:
			meta['opt_11'] = excel.opt_11
		if excel.opt_12:
			meta['opt_12'] = excel.opt_12
		if excel.opt_13:
			meta['opt_13'] = excel.opt_13
		if excel.opt_14:
			meta['opt_14'] = excel.opt_14
		if excel.opt_15:
			meta['opt_15'] = excel.opt_15
		if excel.opt_16:
			meta['opt_16'] = excel.opt_16
		if excel.opt_17:
			meta['opt_17'] = excel.opt_17
		if excel.opt_18:
			meta['opt_18'] = excel.opt_18
		if excel.opt_19:
			meta['opt_19'] = excel.opt_19
		if excel.opt_20:
			meta['opt_20'] = excel.opt_20
		
		if excel.filter_one:
			meta['filter_one'] = excel.filter_one
		if excel.filter_two:
			meta['filter_two'] = excel.filter_two
		if excel.filter_three:
			meta['filter_three'] = excel.filter_three
		if excel.filter_four:
			meta['filter_four'] = excel.filter_four
		if excel.filter_five:
			meta['filter_five'] = excel.filter_five


		object_excel = Excel(
			full_path,
			**meta, 
		)

		object_excel.create_list_pyexcel()
		for position in object_excel.common_list:
			code = position[object_excel.name_column['code']]
			
			good = Goods.objects.filter(code=code).exists()
			
			if good:
				good = Goods.objects.filter(code=code)[0]
			if not good:
				if not excel.only_old:
					print('new', position[object_excel.name_column['code']])
					new_good = Goods()
					new_good.code = code
					new_good.articul = position[object_excel.name_column['articul']]
					check_producer = Producers.objects.filter(name=position[object_excel.name_column['producer']]).exists()
					if check_producer: 
						new_good.producer = Producers.objects.get(name=position[object_excel.name_column['producer']])
					else:
						new_producer = create_producer(position[object_excel.name_column['producer']]) 
						new_good.producer = new_producer
					new_good.description = position[object_excel.name_column['description']]
					if excel.stock:
						new_good.in_stock = position[object_excel.name_column['stock']]
					if excel.price:
						new_good.price = round(position[object_excel.name_column['price']], 2)				
					if excel.opt_0:
						new_good.opt_0 = round(position[object_excel.name_column['opt_0']], 2)
					if excel.opt_1:
						new_good.opt_1 = round(position[object_excel.name_column['opt_1']], 2)
					if excel.opt_2:
						new_good.opt_2 = round(position[object_excel.name_column['opt_2']], 2)			
					if excel.opt_3:
						new_good.opt_3 = round(position[object_excel.name_column['opt_3']], 2)
					if excel.opt_4:
						new_good.opt_4 = round(position[object_excel.name_column['opt_4']], 2)				
					if excel.opt_5:
						new_good.opt_5 = round(position[object_excel.name_column['opt_5']], 2)			
					if excel.opt_6:
						new_good.opt_6 = round(position[object_excel.name_column['opt_6']], 2)		
					if excel.opt_7:
						new_good.opt_7 = round(position[object_excel.name_column['opt_7']], 2)				
					if excel.opt_8:
						new_good.opt_8 = round(position[object_excel.name_column['opt_8']], 2)
					if excel.opt_9:
						new_good.opt_9 = round(position[object_excel.name_column['opt_9']], 2)
					if excel.opt_10:
						new_good.opt_10 = round(position[object_excel.name_column['opt_10']], 2)
					if excel.opt_11:
						new_good.opt_11 = round(position[object_excel.name_column['opt_11']], 2)
					if excel.opt_12:
						new_good.opt_12 = round(position[object_excel.name_column['opt_12']], 2)
					if excel.opt_13:
						new_good.opt_13 = round(position[object_excel.name_column['opt_13']], 2)
					if excel.opt_14:
						new_good.opt_14 = round(position[object_excel.name_column['opt_14']], 2)
					if excel.opt_15:
						new_good.opt_15 = round(position[object_excel.name_column['opt_15']], 2)
					if excel.opt_16:
						new_good.opt_16 = round(position[object_excel.name_column['opt_16']], 2)
					if excel.opt_17:
						new_good.opt_17 = round(position[object_excel.name_column['opt_17']], 2)
					if excel.opt_18:
						new_good.opt_18 = round(position[object_excel.name_column['opt_18']], 2)
					if excel.opt_19:
						new_good.opt_19 = round(position[object_excel.name_column['opt_19']], 2)
					if excel.opt_20:
						new_good.opt_20 = round(position[object_excel.name_column['opt_20']], 2)
					
					

					new_good = filter_add(new_good, excel, position, object_excel)



					new_good.save()
					new_position = NewGoods()
					new_position.position = new_good
					new_position.save()
			
			else:
				print('old', position[object_excel.name_column['code']])
				if excel.articul:
					good.articul = position[object_excel.name_column['articul']] 
				if excel.description:
					good.description = position[object_excel.name_column['description']]
				if excel.stock:
					good.in_stock = position[object_excel.name_column['stock']]
				if excel.price:
					good.price = round(position[object_excel.name_column['price']], 2)
				if excel.opt_0:
					good.opt_0 = round(position[object_excel.name_column['opt_0']], 2)
				if excel.opt_1:
					good.opt_1 = round(position[object_excel.name_column['opt_1']], 2)
				if excel.opt_2:
					good.opt_2 = round(position[object_excel.name_column['opt_2']], 2)
				if excel.opt_3:
					good.opt_3 = round(position[object_excel.name_column['opt_3']], 2)			
				if excel.opt_4:
					good.opt_4 = round(position[object_excel.name_column['opt_4']], 2)
				if excel.opt_5:
					good.opt_5 = round(position[object_excel.name_column['opt_5']], 2)				
				if excel.opt_6:
					good.opt_6 = round(position[object_excel.name_column['opt_6']], 2)
				if excel.opt_7:
					good.opt_7 = round(position[object_excel.name_column['opt_7']], 2)
				if excel.opt_8:
					good.opt_8 = round(position[object_excel.name_column['opt_8']], 2)			
				if excel.opt_9:
					good.opt_9 = round(position[object_excel.name_column['opt_9']], 2)			
				if excel.opt_10:
					good.opt_10 = round(position[object_excel.name_column['opt_10']], 2)
				if excel.opt_11:
					good.opt_11 = round(position[object_excel.name_column['opt_11']], 2)
				if excel.opt_12:
					good.opt_12 = round(position[object_excel.name_column['opt_12']], 2)
				if excel.opt_13:
					good.opt_13 = round(position[object_excel.name_column['opt_13']], 2)
				if excel.opt_14:
					good.opt_14 = round(position[object_excel.name_column['opt_14']], 2)
				if excel.opt_15:
					good.opt_15 = round(position[object_excel.name_column['opt_15']], 2)
				if excel.opt_16:
					good.opt_16 = round(position[object_excel.name_column['opt_16']], 2)
				if excel.opt_17:
					good.opt_17 = round(position[object_excel.name_column['opt_17']], 2)
				if excel.opt_18:
					good.opt_18 = round(position[object_excel.name_column['opt_18']], 2)
				if excel.opt_19:
					good.opt_19 = round(position[object_excel.name_column['opt_19']], 2)
				if excel.opt_20:
					good.opt_20 = round(position[object_excel.name_column['opt_20']], 2)
				
				good = filter_add(good, excel, position, object_excel)	
				good.save()
				
		
		excel.check = True
		excel.save()
	else:
		print('нет новых экселей')
	# except Exception as err:
	# 	print('ОШИБКА ПРИ ОБРАБОТКИ ЭКСЕЛЯ' ,err)
		

	response = 'hello'
	return HttpResponse(response)


def filter_add(good, excel, position, object_excel):
	if excel.filter_one:
		excel_one = position[object_excel.name_column['filter_one']].lower().strip()
		if excel_one:
			db_one = LvlOne.objects.filter(name=excel_one)
			check_category = db_one.exists()
			if check_category:
				good.filter_lvl_one = db_one.first()
			else:
				new_lvl = create_filter_lvl(
					excel_one, 
					LvlOne,
					)
				good.filter_lvl_one = new_lvl

	if excel.filter_two:
		excel_two = position[object_excel.name_column['filter_two']].lower().strip()
		if excel_two:
			db_two = LvlTwo.objects.filter(name=excel_two)
			check_category = db_two.exists()
			if check_category:
				good.filter_lvl_two = db_two.first()
			else:
				new_lvl = create_filter_lvl(
					excel_two, 
					LvlTwo,
					)
				good.filter_lvl_two = new_lvl
	
	if excel.filter_three:
		excel_three = position[object_excel.name_column['filter_three']].lower().strip()
		if excel_three:
			db_three = LvlThree.objects.filter(name=excel_three)
			check_category = db_three.exists()
			if check_category:
				good.filter_lvl_three = db_three.first()
			else:
				new_lvl = create_filter_lvl(
					excel_three, 
					LvlThree,
					)
				good.filter_lvl_three = new_lvl
	
	if excel.filter_four:
		excel_four = position[object_excel.name_column['filter_four']].lower().strip()
		if excel_four:
			db_four = LvlFour.objects.filter(name=excel_four)
			check_category = db_four.exists()
			if check_category:
				good.filter_lvl_four = db_four.first()
			else:
				new_lvl = create_filter_lvl(
					excel_four, 
					LvlFour,
					)
				good.filter_lvl_four = new_lvl
	
	if excel.filter_five:
		excel_five = position[object_excel.name_column['filter_five']].lower().strip()
		if excel_five:
			db_five = LvlFive.objects.filter(name=excel_five)
			check_category = db_five.exists()
			if check_category:
				good.filter_lvl_five = db_five.first()
			else:
				new_lvl = create_filter_lvl(
					excel_five, 
					LvlFive,
					)
				good.filter_lvl_five = new_lvl

	return good




def create_filter_lvl(name, lvl):
	new_lvl = lvl()
	new_lvl.name = name
	new_lvl.save()
	return new_lvl


def create_producer(name):
	new_producer = Producers()
	new_producer.name = name
	new_producer.save()
	return new_producer




def smtp_send(smtp_host, smtp_port, smtp_login, smtp_password, send_to, message_text):
	msg = MIMEText(message_text, 'plain', 'utf-8')
	msg['Subject'] = Header("Ошибка опт-тулс", 'utf-8')
	msg['From'] = smtp_login
	msg["To"] = send_to
	smtpObj = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=10)
	smtpObj.login(smtp_login, smtp_password)
	smtpObj.sendmail(smtp_login, send_to, msg.as_string())
	smtpObj.quit()




# def check_db(code):
# 	check = Goods.objects.filter(code=code).exists()
# 	if check:
# 		new_good = Goods()
# 		new_good.code = code
# 		new_good.articul = 
# 	else:





def opt_price(good, opt):
	price = None
	opt = str(opt)
	if opt == '0':
		price = good.opt_0
	elif opt == '1':
		price = good.opt_1
	elif opt == '2':
		price = good.opt_2
	elif opt == '3':
		price = good.opt_3
	elif opt == '4':
		price = good.opt_4
	elif opt == '5':
		price = good.opt_5
	elif opt == '6':
		price = good.opt_6
	elif opt == '7':
		price = good.opt_7
	elif opt == '8':
		price = good.opt_8
	elif opt == '9':
		price = good.opt_9
	elif opt == '10':
		price = good.opt_10
	elif opt == '11':
		price = good.opt_11
	elif opt == '12':
		price = good.opt_12
	elif opt == '13':
		price = good.opt_13
	elif opt == '14':
		price = good.opt_14
	elif opt == '15':
		price = good.opt_15
	elif opt == '16':
		price = good.opt_16
	elif opt == '17':
		price = good.opt_17
	elif opt == '18':
		price = good.opt_18
	elif opt == '19':
		price = good.opt_19
	elif opt == '20':
		price = good.opt_20
	else:
		price = good.price
	return price




def create_excel(opt_user):
	goods = Goods.objects.all()
	array_goods = []
	excel_header = [
		"Код", 
		"Артикул", 
		"Производитель", 
		"Описание",
		"Наличие",
		"Цена",
		"Оптовая-цена"
		]
	array_goods.append(excel_header)
	for good in goods:
		price_opt = opt_price(good, opt_user)
		row = [
			good.code, 
			good.articul,
			good.producer.name,
			good.description,
			good.in_stock,
			good.price,
			price_opt
			]
		array_goods.append(row)

	
	name = "{}.xlsx".format(random.randint(0, 55000))
	base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
	path = base_dir + '/uploads/generator_excel/' + name
	pyexcel.save_as(array=array_goods, dest_file_name=path)
	return path







class ExcelDownload(View):
	def get(self, request, *args, **kwargs):
		template = get_template('main_page/dowanload_excel_dummy.html')
		opt_user = request.user.customer_set.first().opt
		excel = create_excel(opt_user)	
		context = {
			"data": 'data',
		}
		html = template.render(context)
		excel_content = open(excel, 'rb')
		response = HttpResponse(excel_content.read(), content_type='application/vnd.ms-excel')
		filename = "Invoice_%s.xlsx" %("12341231")
		content = "inline; filename='%s'" %(filename)
		download = request.GET.get("download")
		if download:
			content = "attachment; filename='%s'" %(excel)
		response['Content-Disposition'] = 'inline; filename=' + 'prime-tools.xlsx'
		os.remove(excel)
		return response
		



